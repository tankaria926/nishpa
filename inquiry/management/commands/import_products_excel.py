import os
from django.core.management.base import BaseCommand, CommandError


class Command(BaseCommand):
    help = 'Import products and categories from an Excel file. If no file provided, uses hardcoded data.'

    def add_arguments(self, parser):
        parser.add_argument('excel_path', nargs='?', help='Path to Excel file (.xlsx)')

    def handle(self, *args, **options):
        excel_path = options.get('excel_path')

        # Hardcoded fallback data
        hardcoded = [
            {'category': 'Electronics', 'name': 'Wireless Headphones', 'image_url': 'https://via.placeholder.com/300x200?text=Headphones', 'sku': 'ELE-001'},
            {'category': 'Electronics', 'name': 'Portable Speaker', 'image_url': 'https://via.placeholder.com/300x200?text=Speaker', 'sku': 'ELE-002'},
            {'category': 'Office Supplies', 'name': 'Notebook A5', 'image_url': 'https://via.placeholder.com/300x200?text=Notebook', 'sku': 'OFF-001'},
            {'category': 'Office Supplies', 'name': 'Ballpoint Pen', 'image_url': 'https://via.placeholder.com/300x200?text=Pen', 'sku': 'OFF-002'},
            {'category': 'Furniture', 'name': 'Ergonomic Chair', 'image_url': 'https://via.placeholder.com/300x200?text=Chair', 'sku': 'FUR-001'},
        ]

        rows = []
        if excel_path:
            if not os.path.exists(excel_path):
                raise CommandError(f'Excel file not found: {excel_path}')
            try:
                from openpyxl import load_workbook
            except ImportError:
                raise CommandError('openpyxl is required to import Excel files. Install with pip install openpyxl')

            wb = load_workbook(excel_path, read_only=True)
            ws = wb.active
            headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
            # Expect headers like Category, ProductName, ImageURL, SKU
            mapping = {h: i for i, h in enumerate(headers)}
            for row in ws.iter_rows(min_row=2, values_only=True):
                cat = row[mapping.get('Category')]
                name = row[mapping.get('ProductName')]
                image = row[mapping.get('ImageURL')] if mapping.get('ImageURL') is not None else ''
                sku = row[mapping.get('SKU')] if mapping.get('SKU') is not None else ''
                if cat and name:
                    rows.append({'category': str(cat), 'name': str(name), 'image_url': str(image or ''), 'sku': str(sku or '')})
        else:
            rows = hardcoded

        from inquiry.models import ProductCategory, Product

        created = 0
        for r in rows:
            cat, _ = ProductCategory.objects.get_or_create(name=r['category'])
            p, is_new = Product.objects.get_or_create(
                category=cat,
                name=r['name'],
                defaults={'image_url': r.get('image_url', ''), 'sku': r.get('sku', '')}
            )
            if not is_new:
                p.image_url = r.get('image_url', '')
                p.sku = r.get('sku', '')
                p.save()
            else:
                created += 1

        self.stdout.write(self.style.SUCCESS(f'Imported/updated products (new: {created}).'))
